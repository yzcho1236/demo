import operator
import traceback
import functools
import urllib
from io import BytesIO
from django.contrib.auth import authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Q, AutoField, NOT_PROVIDED
from django.http import HttpResponse, HttpResponseRedirect
from django.utils.decorators import method_decorator
from django.utils.encoding import smart_str, force_str
from django.views import View
from django.contrib import auth
from django.template.response import TemplateResponse
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from demo.form import UserForm, ItemForm, ItemAddForm, RoleForm, RoleAddForm, \
    UserPwdForm, UserEditForm, RegisterForm
from demo.util import Pagination, _filter_map_jqgrid_django, perm_required, select_op
from input.models import Role, User, UserRole, RolePermission, Item, Perm


class WelcomePage(View):
    def get(self, request):
        return TemplateResponse(request, "welcome_page.html")


class Login(View):
    def post(self, request, *args, **kwargs):
        form = UserForm(request.POST)  # POST请求提交表单，该视图将再次创建一个表单实例，数据绑定到表单，包含用户名和密码
        next_url = request.POST.get("next_url", None)
        if next_url == "None":
            next_url = None
        username = request.POST.get('username', None)
        password = request.POST.get('password', None)
        if form.is_valid():
            # 获取表单数据
            try:
                user = User.objects.get(username=username)
            except:
                return TemplateResponse(request, "login.html",
                                        {"username": username, "password": password, "error": "用户名错误"})
            user = authenticate(request, username=username, password=password)
            if user:
                auth.login(request, user)
                # 将用户权限写入session中
                roles = UserRole.objects.filter(user=request.user.id).values_list('role_id')
                perms = list(
                    RolePermission.objects.filter(role_id__in=roles).values_list("permission__codename", flat=True))
                all_perms = list(Perm.objects.all().values_list("codename", flat=True))
                if request.user.is_superuser and request.user.is_active:
                    request.session["perm"] = all_perms
                else:
                    request.session["perm"] = perms

                # 判断是否是重定向
                if next_url:
                    return HttpResponseRedirect(next_url)
                return HttpResponseRedirect("/index/")
            else:
                return TemplateResponse(request, "login.html",
                                        {"username": username, "password": password, "error": "密码错误"})
        else:
            return TemplateResponse(request, "login.html",
                                    {"username": username, "password": password, "error": "登录失败"})

    def get(self, request, *args, **kwargs):
        next_url = request.GET.get("next", None)
        return TemplateResponse(request, "login.html", {"next_url": next_url})


class LogoutView(View):
    """
    用户登出
    """

    def get(self, request):
        auth.logout(request)
        # logout_then_login(request)
        return HttpResponseRedirect('/')


class Register(View):
    def post(self, request):
        form = RegisterForm(request.POST)  # 包含用户名和密码

        # 获取表单数据
        username = request.POST.get('username', None)
        password1 = request.POST.get('password1', None)
        password2 = request.POST.get('password2', None)
        data = {
            "username": username,
            "password1": password1,
            "password2": password2,
            "error": ""
        }
        if form.is_valid():
            # 添加到数据库
            user = User.objects.filter(username=username)
            if user:
                data["error"] = "用户名已存在"
                return TemplateResponse(request, "register.html", data)
            if password1 and password2:
                if password1 != password2:
                    data["error"] = "密码输入不一致，请重新输入密码"
                    return TemplateResponse(request, "register.html", data)
            try:
                User.objects.create_user(username=username, password=password1)
            except:
                data["error"] = "注册失败"
                return TemplateResponse(request, "register.html", data)
        else:
            data["error"] = "注册失败"
            return TemplateResponse(request, "register.html", data)
        return HttpResponseRedirect('/index/')

    def get(self, request, *args, **kwargs):
        return TemplateResponse(request, "register.html")


def index(request):
    perm = request.perm
    # query = request.session.get("query_data", None) if request.session.get("query_data", None) else []
    return TemplateResponse(request, "index.html", {"perm": perm})


class ItemView(View):
    """查看物料列表"""
    file_headers = ('id', '代码', '名称', '条码')
    select_field = {'id': 'id', 'nr': '代码', 'name': '名称', 'barcode': '条码'}

    @method_decorator(login_required)
    @method_decorator(perm_required(("view_item",)))
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('format', None)
        if fmt is None:
            data = self._get_data(request)
            return TemplateResponse(request, "item.html", data)

        elif fmt == 'spreadsheet':
            # 下载 excel
            json = self._get_data(request, in_page=False)
            wb = Workbook(write_only=True)
            title = "物料"
            ws = wb.create_sheet(title)

            headers = []
            for h in self.file_headers:
                cell = WriteOnlyCell(ws, value=h)
                headers.append(cell)
            ws.append(headers)

            body_fields = ("id", "nr", "name", "barcode")
            for data in json['data']:
                body = []
                for field in body_fields:
                    if field in data:
                        cell = WriteOnlyCell(ws, value=data[field])
                        body.append(cell)
                ws.append(body)

            output = BytesIO()
            wb.save(output)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                content=output.getvalue()
            )
            response['Content-Disposition'] = "attachment; filename*=utf-8''%s.xlsx" % urllib.parse.quote(
                force_str(title))
            response['Cache-Control'] = "no-cache, no-store"
            return response

    @method_decorator(login_required)
    @method_decorator(perm_required(("add_item",)))
    def post(self, request, *args, **kwargs):
        """上传Excel"""
        if request.FILES and len(request.FILES) == 1:
            count = 0
            for filename, file in request.FILES.items():
                if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    count += 1
            if count == 0:
                data = self._get_data(request)
                data["error"] = "没有上传文件"
                return TemplateResponse(request, "item.html", data)

            else:
                for name, file in request.FILES.items():
                    wb = load_workbook(filename=file, read_only=True, data_only=True)
                    # 第一个sheet
                    sheet = wb[wb.sheetnames[0]]
                    row_count = 0
                    row_number = 1
                    # 对应表头是在第几列
                    headers_index = {}
                    # Excel传入的字段和模型类字段的对应关系
                    headers_field_name = {}

                    for row in sheet.iter_rows():
                        row_count += 1
                        if row_number == 1:
                            row_number += 1
                            # 获取excelheader
                            _headers = [i.value for i in row]
                            index = 0
                            for h in _headers:
                                for field in Item._meta.fields:
                                    if h == field.verbose_name:
                                        headers_index[h] = index
                                        headers_field_name[h] = field.name
                                index += 1
                        # 取值
                        else:
                            values = [i.value for i in row]
                            none_len = 0
                            for v in values:
                                if v is None:
                                    none_len += 1

                            if none_len == len(values):
                                continue
                            upload_fields = [v for k, v in headers_field_name.items()]

                            # 必传字段
                            required_fields = set()
                            for i in Item._meta.fields:
                                if not i.blank and i.default == NOT_PROVIDED and not isinstance(i, AutoField):
                                    required_fields.add(i.name)
                            set_header = set(upload_fields)

                            set_header.discard("id")
                            if set_header != required_fields:
                                data = self._get_data(request)
                                data["error"] = "上传字段不全，请重新上传"
                                return TemplateResponse(request, "item.html", data)

                            # 有id值进行更新或者创建，没有id值进行创建
                            if headers_index["id"] is not None and values[headers_index["id"]]:
                                dict = {}
                                item_query = Item.objects.filter(id=values[headers_index["id"]]).first()
                                for k, v in headers_index.items():
                                    value = values[v]
                                    field_name = headers_field_name[k]
                                    # 有id值更新
                                    if item_query:
                                        if field_name == "id":
                                            item_query.id = value
                                        if field_name == 'nr' and value is not None:
                                            item_query.nr = value
                                        elif field_name == 'name' and value is not None:
                                            item_query.name = value
                                        elif field_name == 'barcode' and value is not None:
                                            item_query.barcode = value
                                    # 有id，是无效的id值创建
                                    else:
                                        if field_name == "id":
                                            continue
                                        if field_name == 'nr' and value is not None:
                                            dict["nr"] = value
                                        elif field_name == 'name' and value is not None:
                                            dict["name"] = value
                                        elif field_name == 'barcode' and value is not None:
                                            dict["barcode"] = value
                                if not item_query:
                                    if len(dict) != len(self.file_headers) - 1:
                                        data = self._get_data(request)
                                        data["error"] = "字段值不能为空"
                                        return TemplateResponse(request, "item.html", data)
                                    try:
                                        Item.objects.create(nr=dict["nr"], name=dict["name"], barcode=dict["barcode"])
                                    except Exception as e:
                                        data = self._get_data(request)
                                        data["error"] = "数据已存在，无法上传"
                                        return TemplateResponse(request, "item.html", data)
                                else:
                                    item_query.save()

                            # 无id值创建
                            else:
                                dict = {}
                                for k, v in headers_index.items():
                                    value = values[v]
                                    field_name = headers_field_name[k]
                                    if field_name == "id":
                                        continue
                                    if field_name == 'nr' and value is not None:
                                        dict["nr"] = value
                                    elif field_name == 'name' and value is not None:
                                        dict["name"] = value
                                    elif field_name == 'barcode' and value is not None:
                                        dict["barcode"] = value

                                if len(dict) != len(self.file_headers) - 1:
                                    data = self._get_data(request)
                                    data["error"] = "字段值不能为空"
                                    return TemplateResponse(request, "item.html", data)

                                try:
                                    Item.objects.create(nr=dict["nr"], name=dict["name"], barcode=dict["barcode"])
                                except Exception as e:
                                    data = self._get_data(request)
                                    data["error"] = "数据已存在，无法上传"
                                    return TemplateResponse(request, "item.html", data)

                data = self._get_data(request)
                return TemplateResponse(request, "item.html", data)

        else:
            data = self._get_data(request)
            data["error"] = "没有上传文件"
            return TemplateResponse(request, "item.html", data)

    def _get_data(self, request, in_page=True, *args, **kwargs):
        # 查询的条件
        q_filters = []
        # 查询的字段
        filter_fields = ("id", "nr", "name", "barcode")
        # 获取到的页面数
        page = request.GET.get("page", 1)
        query_data = dict(request.GET.lists())
        fs = []
        if "data" in query_data and "op" in query_data and "data" in query_data:
            data_query = list(filter(lambda x: x, query_data["data"]))
            len_data_query = len(data_query)
            fields = query_data["field"]
            ops = query_data["op"]
            # 查询的有效列表
            # 查询的所有列表
            for i in range(0, len_data_query):
                adict = {}
                if fields[i] in filter_fields:
                    adict["field"] = fields[i]
                    adict["op"] = ops[i]
                    adict["data"] = data_query[i]
                    fs.append(adict)
                else:
                    continue

            for rule in fs:
                op, field, data = rule['op'], rule['field'], rule['data']
                filter_fmt, exclude = _filter_map_jqgrid_django[op]
                filter_str = smart_str(filter_fmt % {'field': field})

                if filter_fmt.endswith('__in'):
                    filter_kwargs = {filter_str: data.split(',')}
                else:
                    filter_kwargs = {filter_str: smart_str(data)}

                if exclude:
                    q_filters.append(~Q(**filter_kwargs))
                else:
                    q_filters.append(Q(**filter_kwargs))

        if q_filters:
            item_query = Item.objects.filter(functools.reduce(operator.iand, q_filters)).order_by("id")
        else:
            item_query = Item.objects.all().order_by("id")
        count = float(item_query.count())
        pagesize = request.pagesizes if in_page else count
        pagination = Pagination(item_query, pagesize, page)
        content = []
        for i in pagination.get_objs():
            data = {
                "id": i.id,
                "nr": i.nr,
                "name": i.name,
                "barcode": i.barcode,
            }
            content.append(data)

        # 根据查询的列表拼接URL
        array = []
        for i in fs:
            for k, v in i.items():
                b = "&" + str(k) + "=" + str(v)
                array.append(b)
        query_url = "".join(array)

        all_data = {"page": pagination.page, "s": pagination.count, "total_pages": pagination.total_pages,
                    "data": content, "pg": pagination,
                    "perm": request.perm, "query": fs,
                    'query_url': query_url, "select_field": self.select_field,
                    "select_op": select_op, "error": ""}

        return all_data


class ItemEdit(LoginRequiredMixin, View):
    permission_required = "change_item"

    def get(self, request, *args, **kwargs):
        item_id = request.GET.get('item_id', None)
        try:
            item = Item.objects.get(id=item_id)
            data = {
                "id": item.id,
                "nr": item.nr,
                "name": item.name,
                "barcode": item.barcode,
            }
        except:
            return HttpResponseRedirect("/item/?msg=数据查询失败")
        return TemplateResponse(request, "item_edit.html", {"data": data, "perm": request.perm})

    def post(self, request, *args, **kwargs):
        form = ItemForm(request.POST)
        # 获取表单数据
        item_id = request.POST['id']
        nr = request.POST['nr']
        name = request.POST['name']
        barcode = request.POST['barcode']
        data = {
            "id": item_id,
            "nr": nr,
            "name": name,
            "barcode": barcode,

        }
        if form.is_valid():

            with transaction.atomic(savepoint=False):
                # 创建保存点
                try:
                    item = Item.objects.get(id=item_id)
                    item.nr = nr
                    item.name = name
                    item.barcode = barcode
                    item.save()
                except:
                    return TemplateResponse(request, "item_edit.html",
                                            {"data": data, "perm": request.perm, "error": "编辑失败"})
                else:
                    return HttpResponseRedirect("/item/")
        else:
            return TemplateResponse(request, "item_edit.html", {"data": data, "perm": request.perm, "error": "表单填写错误"})


class ItemDelete(LoginRequiredMixin, View):
    permission_required = "delete_item"

    def get(self, request, *args, **kwargs):
        item_id = request.GET.get('item_id', None)
        try:
            item = Item.objects.get(id=item_id)
            data = {
                "id": item.id,
                "nr": item.nr,
                "name": item.name,
                "barcode": item.barcode,
                "perm": request.perm
            }
        except Exception as e:
            traceback.print_exc()
            return HttpResponseRedirect("/item/?msg=数据查询失败")

        return TemplateResponse(request, "item_delete.html", {"data": data})

    def post(self, request, *args, **kwargs):
        item_id = request.POST['id']
        nr = request.POST['nr']
        name = request.POST['name']
        barcode = request.POST['barcode']
        data = {
            "id": item_id,
            "nr": nr,
            "name": name,
            "barcode": barcode,
        }

        with transaction.atomic(savepoint=False):
            # 创建保存点
            try:
                item = Item.objects.get(id=item_id)
                item.delete()
            except:
                return TemplateResponse(request, "item_delete.html", {"data": data, "error": "编辑失败"})

            else:
                return HttpResponseRedirect('/item/')


class ItemAdd(LoginRequiredMixin, View):
    permission_required = "add_item"

    def get(self, request, *args, **kwargs):

        return TemplateResponse(request, "item_add.html", {"perm": request.perm})

    def post(self, request, *args, **kwargs):
        form = ItemAddForm(request.POST)
        # 获取表单数据
        nr = request.POST['nr']
        name = request.POST['name']
        barcode = request.POST['barcode']
        data = {
            "nr": nr,
            "name": name,
            "barcode": barcode
        }
        if form.is_valid():

            with transaction.atomic(savepoint=False):
                try:
                    Item.objects.create(nr=nr, name=name, barcode=barcode)
                except:
                    return TemplateResponse(request, "item_add.html", {"data": data, "error": "添加失败，物料代码已存在"})
                else:
                    return HttpResponseRedirect("/item/")
        else:
            return TemplateResponse(request, "item_edit.html", {"data": data, "error": "表单错误"})


class UserView(LoginRequiredMixin, View):
    permission_required = "view_user"

    def get(self, request, *args, **kwargs):
        """获取用户列表"""

        content = []
        page = request.GET.get("page", 1)
        pagesize = request.pagesizes
        user_query = User.objects.all().order_by("id")
        pagination = Pagination(user_query, pagesize, page)

        for i in pagination.get_objs():
            data = {
                "id": i.id,
                "username": i.username,
                "is_superuser": i.is_superuser,
            }
            content.append(data)
        all_data = {"page": page, "records": pagination.count, "total_pages": pagination.total_pages, "data": content,
                    "perm": request.perm}

        return TemplateResponse(request, "user.html", all_data)


class UserEditInfo(LoginRequiredMixin, View):
    permission_required = "change_user"

    def get(self, request, *args, **kwargs):
        user_id = request.GET.get('id', None)
        try:
            user = User.objects.get(id=user_id)
            data = {
                "id": user.id,
                "username": user.username,
                "is_superuser": user.is_superuser,
                "perm": request.perm
            }

        except:
            return HttpResponseRedirect("/user/?msg=数据查询失败")

        return TemplateResponse(request, "user_edit.html", data)

    def post(self, request, *args, **kwargs):
        form = UserEditForm(request.POST)
        # 获取表单数据
        id = request.POST.get('id')
        username = request.POST.get('username')
        is_superuser = request.POST.get('is_superuser')
        data = {
            "error": "",
            "id": id,
            "username": username,
            "is_superuser": is_superuser,
            "perm": request.perm
        }

        if form.is_valid():
            with transaction.atomic(savepoint=False):
                # 创建保存点
                try:
                    user = User.objects.get(id=id)
                    user.username = username
                    user.is_superuser = is_superuser
                    user.save()
                    # 用户更改密码后更新用户session设置，更改用户密码后不进行自动登出
                    update_session_auth_hash(request, user)
                except:
                    data["error"] = "用户编辑失效"
                    return TemplateResponse(request, "user_edit.html", data)
                else:
                    return HttpResponseRedirect("/user/")
        else:
            data["error"] = "表单填写错误"
            data["form"] = form
            return TemplateResponse(request, "user_edit.html", data)


class UserEditPwd(LoginRequiredMixin, View):
    permission_required = "change_user"

    def get(self, request, *args, **kwargs):
        id = request.GET.get('id', None)
        try:
            user = User.objects.get(id=id)
            data = {
                "id": user.id,
                "username": user.username,
                "password1": "",
                "password2": "",
                "perm": request.perm
            }

        except:
            return HttpResponseRedirect("/user/?msg=数据查询失败")

        return TemplateResponse(request, "user_edit_password.html", data)

    def post(self, request, *args, **kwargs):
        form = UserPwdForm(request.POST)
        # 获取表单数据
        id = request.POST.get('id')
        username = request.POST.get('username')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        data = {
            "error": "",
            "id": id,
            "username": username,
            "password1": password1,
            "password2": password2,
            "perm": request.perm
        }

        if form.is_valid():
            if password1 and password2:
                if password1 != password2:
                    data["error"] = "密码输入不一致，请重新输入密码"
                    return TemplateResponse(request, "user_edit_password.html", data)

            with transaction.atomic(savepoint=False):
                # 创建保存点
                try:
                    user = User.objects.get(id=id)
                    user.set_password(password1)
                    user.save()
                    # 用户更改密码后更新用户session设置，更改用户密码后不进行自动登出
                    # update_session_auth_hash(request, user)
                except:
                    data["error"] = "用户编辑失效"
                    return TemplateResponse(request, "user_edit_password.html", data)
                else:
                    return HttpResponseRedirect("/user/")
        else:
            data["error"] = "表单填写错误"
            data["form"] = form
            return TemplateResponse(request, "user_edit_password.html", data)


class UserDelete(LoginRequiredMixin, View):
    permission_required = "delete_user"

    def get(self, request, *args, **kwargs):
        user_id = request.GET.get('user_id', None)
        try:
            user = User.objects.get(id=user_id)
            data = {
                "id": user.id,
                "username": user.username,
                "is_superuser": user.is_superuser,
                "perm": request.perm
            }
        except:
            return HttpResponseRedirect("/user/?msg=数据查询失败")

        return TemplateResponse(request, "user_delete.html", {"data": data})

    def post(self, request, *args, **kwargs):
        user_id = request.POST.get('id')
        username = request.POST.get('username')
        is_superuser = request.POST.get('is_superuser')
        password = request.POST.get('password')
        data = {
            "error": "",
            "id": user_id,
            "username": username,
            "is_superuser": is_superuser,
            "perm": request.perm
        }

        with transaction.atomic(savepoint=False):
            # 创建保存点
            try:
                user = User.objects.get(id=user_id)
                user.delete()
            except Exception as e:
                data["error"] = "删除用户失败"
                return TemplateResponse(request, "user_delete.html", data)
            else:
                return HttpResponseRedirect('/user/')


class RoleView(LoginRequiredMixin, View):
    permission_required = "view_role"

    def get(self, request, *args, **kwargs):
        """获取用户列表"""

        content = []
        page = request.GET.get("page", 1)
        pagesize = request.pagesizes
        role_query = Role.objects.all().order_by("id")
        pagination = Pagination(role_query, pagesize, page)

        for i in pagination.get_objs():
            data = {
                "id": i.id,
                "name": i.name,
                "perm": request.perm
            }
            content.append(data)
        all_data = {"page": page, "records": pagination.count, "total_pages": pagination.total_pages, "data": content,
                    "perm": request.perm}

        return TemplateResponse(request, "role.html", all_data)


class RoleEdit(LoginRequiredMixin, View):
    permission_required = "change_role"

    def get(self, request, *args, **kwargs):

        role_id = request.GET.get('role_id', None)
        try:
            role = Role.objects.get(id=role_id)
            data = {
                "id": role.id,
                "name": role.name,
                "perm": request.perm
            }
        except:
            return HttpResponseRedirect("/role/?msg=查询角色信息失败")

        return TemplateResponse(request, "role_edit.html", {"data": data})

    def post(self, request, *args, **kwargs):
        form = RoleForm(request.POST)
        role_id = request.POST['id']
        name = request.POST['name']
        data = {
            "id": role_id,
            "name": name,
            "perm": request.perm
        }
        if form.is_valid():
            # 获取表单数据
            with transaction.atomic(savepoint=False):
                # 创建保存点
                try:
                    role = Role.objects.get(id=role_id)
                    role.name = name
                    role.save()
                except:

                    return TemplateResponse(request, "role_edit.html", {"data": data, "error": "编辑失败"})
                else:
                    return HttpResponseRedirect("/role/")
        else:
            return TemplateResponse(request, "role_edit.html", {"data": data, "error": "表单填写错误"})


class RoleAdd(LoginRequiredMixin, View):
    permission_required = "add_role"

    def get(self, request, *args, **kwargs):

        if request.user.has_perm("add_role"):
            return TemplateResponse(request, "role_add.html")
        else:
            return HttpResponseRedirect("/role/?msg=用户没有添加角色的权限")

    def post(self, request, *args, **kwargs):
        if request.user.has_perm("add_role"):
            form = RoleAddForm(request.POST)
            name = request.POST['name']
            data = {
                "name": name,
                "perm": request.perm
            }

            if form.is_valid():
                # 获取表单数据
                with transaction.atomic(savepoint=False):
                    try:
                        Role.objects.create(name=name)
                    except Exception as e:
                        print(e)
                        return TemplateResponse(request, "role_add.html", {"data": data, "error": "添加失败,角色名称已存在"})
                    else:
                        return HttpResponseRedirect("/role/")
            else:
                return TemplateResponse(request, "role_add.html", {"data": data, "error": "表单错误"})
        else:
            return HttpResponseRedirect("/role/?msg=用户没有添加角色的权限")


class RoleDelete(LoginRequiredMixin, View):
    permission_required = "delete_role"

    def get(self, request, *args, **kwargs):

        role_id = request.GET.get('id', None)
        try:
            role = Role.objects.get(id=role_id)
            data = {
                "id": role.id,
                "name": role.name,
                "perm": request.perm
            }
        except:
            return HttpResponseRedirect("/role/?msg=数据查询失败")

        return TemplateResponse(request, "role_delete.html", {"data": data})

    def post(self, request, *args, **kwargs):
        id = request.POST.get('id', None)
        name = request.POST.get('name', None)
        data = {
            "id": id,
            "name": name,
            "perm": request.perm

        }
        with transaction.atomic(savepoint=False):
            # 创建保存点
            try:
                role = Role.objects.get(id=id)
                role.delete()
            except Exception as e:
                return TemplateResponse(request, "item_delete.html", {"data": data, "error": "编辑失败"})

            else:
                return HttpResponseRedirect('/role/')


class PermissionView(LoginRequiredMixin, View):
    """权限"""
    permission_required = "view_perm"

    def get(self, request, *args, **kwargs):
        content = []
        page = request.GET.get("page", 1)
        pagesize = request.pagesizes
        perm_query = Perm.objects.all().order_by("id")
        pagination = Pagination(perm_query, pagesize, page)
        for i in pagination.get_objs():
            data = {
                "id": i.id,
                "name": i.name,
                "codename": i.codename,
                "perm": request.perm
            }
            content.append(data)

        all_data = {"page": pagination.page, "records": pagination.count, "total_pages": pagination.total_pages,
                    "data": content, "perm": request.perm}

        return TemplateResponse(request, "permission.html", all_data)


class UserRoleView(LoginRequiredMixin, View):
    """用户角色"""

    permission_required = "view_user_role"

    def get(self, request, *args, **kwargs):
        content = {
            "error": "",
            "data": [],
            "all_role_dict": [],
            "perm": request.perm,
        }
        user_role_list = []
        page = request.GET.get("page", 1)
        pagesize = request.pagesizes
        user_query = User.objects.all().order_by("id")
        pagination = Pagination(user_query, pagesize, page)

        all_roles = Role.objects.all().values("id", "name").order_by('id')
        all_role_dict = {}
        for i in list(all_roles):
            all_role_dict[i["id"]] = i["name"]

        for user in pagination.get_objs():
            roles = UserRole.objects.filter(user=user).values("role__name").order_by('id')
            roles_list = [i["role__name"] for i in list(roles)]
            user_role = {
                "id": user.id,
                "user": user.username,
                "role": roles_list,
            }
            user_role_list.append(user_role)
        content["data"] = user_role_list
        content["all_role_dict"] = all_role_dict
        content["page"] = pagination.page
        content["records"] = pagination.count
        content["total_pages"] = pagination.total_pages

        return TemplateResponse(request, "user_role.html", content)


class UserRoleEdit(LoginRequiredMixin, View):
    permission_required = "change_user_role"

    def get(self, request, *args, **kwargs):
        user_id = request.GET.get("user_id", None)
        try:
            user = User.objects.get(id=user_id)
        except:
            return HttpResponseRedirect("/user_role/?msg=没有找到数据")
        roles = UserRole.objects.filter(user=user).values("role__name").order_by('id')
        all_roles = Role.objects.all().values("id", "name").order_by('id')
        all_role_dict = {}
        for i in list(all_roles):
            all_role_dict[i["id"]] = i["name"]

        roles_list = [i["role__name"] for i in list(roles)]
        data = {
            "id": user_id,
            "user": user.username,
            "roles_list": roles_list,
            "all_role_dict": all_role_dict,
            "perm": request.perm
        }
        return TemplateResponse(request, "user_role_edit.html", data)

    def post(self, request, *args, **kwargs):
        user_id = request.POST.get("id")
        user = request.POST.get('user')
        roles = request.POST.get('array')

        if roles:
            role_list = roles.split(',')
            role_int_list = list(map(lambda x: int(x), list(filter(lambda x: type(x) is str, role_list))))
        else:
            role_int_list = []
        with transaction.atomic(savepoint=False):
            try:
                UserRole.objects.filter(user_id=user_id).delete()
                if role_int_list:
                    for i in role_int_list:
                        role = Role.objects.get(id=i)
                        UserRole.objects.create(user=user, role=role)
            except Exception as e:
                return HttpResponseRedirect("/user_role/edit/?msg=编辑用户角色信息失败")

            return HttpResponseRedirect("/user_role/")


class RolePermissionView(LoginRequiredMixin, View):
    """角色权限"""
    permission_required = "view_role_permission"

    def get(self, request, *args, **kwargs):
        content = {
            "error": "",
            "data": [],
            "all_perm_dict": {},
            "perm": request.perm

        }
        page = request.GET.get("page", 1)
        pagesize = request.pagesizes
        roles_query = Role.objects.all().order_by('id')
        pagination = Pagination(roles_query, pagesize, page)

        role_perm_list = []
        all_perms = list(Perm.objects.all().values("id", "codename").order_by('id'))
        all_perm_dict = {}
        for i in all_perms:
            all_perm_dict[i["id"]] = i["codename"]

        # for role in roles:
        #     perms = RolePermission.objects.filter(role=role).values("permission__codename").order_by('id')
        #     perms_list = [i["permission__codename"] for i in list(perms)]
        #     user_role = {
        #         "id": role.id,
        #         "role": role.name,
        #         "perm": perms_list,
        #     }

        perms = list(RolePermission.objects.filter(role_id__in=[i.id for i in list(roles_query)]))
        for role in pagination.get_objs():
            perms_ids = list(map(lambda x: x.permission_id, filter(lambda i: i.role_id == role.id, perms)))
            perms_list = list(
                map(lambda x: x['codename'], filter(lambda i: perms_ids.__contains__(i["id"]), all_perms)))
            user_role = {
                "id": role.id,
                "role": role.name,
                "perm": perms_list,
            }
            role_perm_list.append(user_role)

        content["data"] = role_perm_list
        content["all_perm_dict"] = all_perm_dict
        content["page"] = pagination.page
        content["records"] = pagination.count
        content["total_pages"] = pagination.total_pages
        return TemplateResponse(request, "role_permission.html", content)


class RolePermissionEdit(LoginRequiredMixin, View):
    """角色权限"""
    permission_required = "change_role_permission"

    def get(self, request, *args, **kwargs):
        role_id = request.GET.get("role_id", None)
        try:
            role = Role.objects.get(id=role_id)
        except:
            return HttpResponseRedirect("/role_permission/?msg=没有找到数据")
        perms = RolePermission.objects.filter(role=role).values("permission__codename").order_by('id')
        perms_list = [i["permission__codename"] for i in list(perms)]
        all_perm = Perm.objects.all().values("id", "codename").order_by('id')
        all_perm_dict = {}
        for i in list(all_perm):
            all_perm_dict[i["id"]] = i["codename"]

        role_perm = {
            "id": role.id,
            "role": role.name,
            "perms_list": perms_list,
            "all_perm_dict": all_perm_dict,
            "perm": request.perm
        }

        return TemplateResponse(request, "role_permission_edit.html", role_perm)

    def post(self, request, *args, **kwargs):
        role_id = request.POST["id"]
        permissions = request.POST['array']
        perm_list = permissions.split(',')
        # 判断权限是列表是否为空
        perm_int_list = set(
            map(lambda x: int(x), list(filter(lambda x: type(x) is str, perm_list)))) if permissions else set()

        with transaction.atomic(savepoint=False):
            try:
                pids = set(RolePermission.objects.filter(role_id=role_id).values_list('permission_id', flat=True))
                # 删除被取消的权限
                delete_ids = pids.difference(perm_int_list)
                # 创建新添加的权限
                add_ids = perm_int_list.difference(pids)

                RolePermission.objects.filter(role_id=role_id, permission_id__in=delete_ids).delete()
                RolePermission.objects.bulk_create(
                    [RolePermission(role_id=role_id, permission_id=pid) for pid in add_ids])
            except:
                return HttpResponseRedirect("/role_permission/edit/?msg=修改角色权限信息失败")
            return HttpResponseRedirect("/role_permission/")
