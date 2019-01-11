import json
import logging
import traceback
import urllib
from datetime import datetime
from io import BytesIO
from reportlab.pdfgen import canvas
import requests
from django.contrib.contenttypes.models import ContentType
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import NOT_PROVIDED, AutoField, Q
import uuid
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.utils import timezone
from django.utils.encoding import force_str
from django.views import View, generic
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from django.core.files.uploadedfile import InMemoryUploadedFile
from demo.util import Pagination, select_op
from input.models import Item
from my_app.form import ItemBomForm
from my_app.models import BomModel, Pictures, UploadFileModel
from my_app.qiniu import storage, delete_qiniu_file
from my_app.util import BomNr, BomData
from settings import MEDIA_ROOT, QINIU_DOMIN_PREFIX


class ItemBom(View):
    file_headers = ('id', '物料', '父类', '代码', '生效开始', "生效结束", "数量")
    body_fields = ("id", "item", "parent", "nr", "effective_start", "effective_end", "qty")
    select_field = {"id": "id", 'item__nr': '物料', 'parent__nr': '父类', 'nr': '代码', 'effective_start': '生效开始',
                    'effective_end': '生效结束'}

    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('format', None)
        if fmt is None:
            content = self._get_data(request)
            return TemplateResponse(request, "my_app_template/bom.html", content)

        elif fmt == 'spreadsheet':
            # 下载 excel
            json = self._get_data(request, in_page=False)
            wb = Workbook(write_only=True)
            title = "物料清单"
            ws = wb.create_sheet(title)

            headers = []
            for h in self.file_headers:
                cell = WriteOnlyCell(ws, value=h)
                headers.append(cell)
            ws.append(headers)

            data = []
            for i in json["download"]:
                adict = {}
                adict["id"] = i.id
                adict["item"] = i.item.nr
                if i.parent:
                    adict["parent"] = i.parent.nr
                else:
                    adict["parent"] = ""
                adict["nr"] = i.nr
                adict["effective_start"] = i.effective_start
                adict["effective_end"] = i.effective_end
                adict["qty"] = i.qty
                data.append(adict)

            for i in data:
                body = []
                for field in self.body_fields:
                    if field in i:
                        cell = WriteOnlyCell(ws, value=i[field])
                        body.append(cell)
                ws.append(body)
            output = BytesIO()
            wb.save(output)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                content=output.getvalue()
            )
            response['Content-Disposition'] = "attachment; filename*=utf-8''%s.xlsx" % urllib.parse.quote(
                force_str(title))
            response['Cache-Control'] = "no-cache, no-store"
            return response

    def post(self, request, *args, **kwargs):
        # 上传Excel
        data = self._get_data(request)
        if request.FILES and len(request.FILES) == 1:
            count = 0
            for filename, file in request.FILES.items():
                if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    count += 1
            if count == 0:
                data = self._get_data(request)
                data["error"] = "没有上传文件"
                return TemplateResponse(request, "my_app_template/bom.html", data)
            else:
                for name, file in request.FILES.items():
                    wb = load_workbook(filename=file, read_only=True, data_only=True)
                    # 第一个sheet
                    sheet = wb[wb.sheetnames[0]]
                    row_count = 0
                    row_number = 1
                    # 对应表头是在第几列
                    headers_index = {}
                    # Excel传入的字段和模型类字段的对应关系
                    headers_field_name = {}
                    excel_level = 0
                    for row in sheet.iter_rows():
                        excel_level += 1
                        row_count += 1
                        if row_number == 1:
                            row_number += 1
                            # 获取excelheader
                            _headers = [i.value for i in row]
                            index = 0
                            for h in _headers:
                                for field in BomModel._meta.fields:
                                    if h == field.verbose_name:
                                        headers_index[h] = index
                                        headers_field_name[h] = field.name
                                index += 1
                        # 取值
                        else:
                            values = [i.value for i in row]
                            none_len = 0
                            for v in values:
                                if v is None:
                                    none_len += 1

                            if none_len == len(values):
                                continue
                            upload_fields = [v for k, v in headers_field_name.items()]

                            # 必传字段
                            required_fields = set()
                            for i in BomModel._meta.fields:
                                if not i.blank and i.default == NOT_PROVIDED and not isinstance(i, AutoField):
                                    required_fields.add(i.name)
                            set_header = set(upload_fields)

                            set_header.discard("id")
                            if set_header.intersection(required_fields) != required_fields:
                                data = self._get_data(request)
                                data["error"] = "上传字段不全，请重新上传"
                                return TemplateResponse(request, "my_app_template/bom.html", data)

                            id = values[headers_index["id"]]
                            nr = values[headers_index["代码"]]
                            item = values[headers_index["物料"]]
                            parent = values[headers_index["父类"]]
                            qty = values[headers_index["数量"]]
                            effective_start = values[headers_index["生效开始"]] if values[
                                headers_index["生效开始"]] else datetime(datetime.now().year, datetime.now().month,
                                                                     datetime.now().day)
                            effective_end = values[headers_index["生效结束"]] if values[
                                headers_index["生效结束"]] else datetime(2030, 12, 31)

                            # 判断物料和父类是否相等
                            if item == parent:
                                data["error"] = "第 %s 行 物料和父类不能为同一个物料" % excel_level
                                return TemplateResponse(request, "my_app_template/bom.html", data)

                            # 判断上传的字段值是否为空
                            upload_value = {"item": item, "qty": qty, "nr": nr}
                            for k, v in upload_value.items():
                                if v is None:
                                    data["error"] = "第 %s 行 %s字段值不能为空" % (excel_level, k)
                                    return TemplateResponse(request, "my_app_template/bom.html", data)

                            # 判断上传的物料代码和父类代码是否存在
                            try:
                                bom_item = Item.objects.get(nr=item)
                            except Item.DoesNotExist as e:
                                traceback.print_exc()
                                data["error"] = "第 %s 行物料为%s数据不存在" % (excel_level, item)
                                return TemplateResponse(request, "my_app_template/bom.html", data)
                            try:
                                bom_parent = Item.objects.get(nr=parent)
                            except Item.DoesNotExist as e:
                                data["error"] = "第 %s 行父类为%s数据不存在" % (excel_level, parent)
                                return TemplateResponse(request, "my_app_template/bom.html", data)

                            # 有id值进行更新或者创建，没有id值进行创建
                            if headers_index["id"] is not None and values[headers_index["id"]]:
                                bom_query = BomModel.objects.filter(id=id)
                                # 有id值更新有效期、数量
                                if bom_query:
                                    bom_detail = BomModel.objects.filter(id=id, nr=nr, item__nr=item,
                                                                         parent__nr=parent).first()
                                    if bom_detail:
                                        bom_detail.qty = qty
                                        bom_detail.effective_start = effective_start
                                        bom_detail.effective_end = effective_end
                                        bom_detail.save()
                                    else:
                                        data["error"] = "第 %s 行数据不存在" % excel_level
                                        return TemplateResponse(request, "my_app_template/bom.html", data)

                                # 有id值，但是为无效的id
                                else:
                                    try:
                                        if parent:
                                            if BomModel.objects.filter(item=bom_parent):
                                                data["error"] = "第 %s 行请先创建父类信息" % excel_level
                                                return TemplateResponse(request, "my_app_template/bom.html", data)

                                            else:
                                                BomModel.objects.create(nr=nr, item=bom_item,
                                                                        parent=bom_parent, qty=qty,
                                                                        effective_start=effective_start,
                                                                        effective_end=effective_end)

                                        else:
                                            BomModel.objects.create(nr=nr, item=bom_item,
                                                                    qty=qty,
                                                                    effective_start=effective_start,
                                                                    effective_end=effective_end)
                                    except Exception as e:
                                        data = self._get_data(request)
                                        data["error"] = "数据已存在，无法上传"
                                        return TemplateResponse(request, "my_app_template/bom.html", data)

                            # 无id值创建
                            else:
                                try:
                                    if parent:
                                        if not BomModel.objects.filter(item=bom_parent):
                                            data["error"] = "第 %s 行请先创建父类信息" % excel_level
                                            return TemplateResponse(request, "my_app_template/bom.html", data)
                                        else:
                                            BomModel.objects.create(nr=nr, item=bom_item,
                                                                    parent=bom_parent, qty=qty,
                                                                    effective_start=effective_start,
                                                                    effective_end=effective_end)
                                    else:
                                        BomModel.objects.create(nr=nr, item=bom_item,
                                                                qty=qty,
                                                                effective_start=effective_start,
                                                                effective_end=effective_end)
                                except Exception as e:
                                    print(e)
                                    traceback.print_exc()
                                    data = self._get_data(request)
                                    data["error"] = "数据已存在，无法上传"
                                    return TemplateResponse(request, "my_app_template/bom.html", data)

                return HttpResponseRedirect("/item/bom/")

        else:
            data["error"] = "没有上传文件"
            return TemplateResponse(request, "my_app_template/bom.html", data)

    def _get_data(self, request, in_page=True, *args, **kwargs):
        page = request.GET.get("page", 1)
        query_data = dict(request.GET.lists())
        bom_query, node_data, download, fs, query_url, search = BomData.get_tree(BomModel, query_data,
                                                                                 self.select_field)
        if search:
            request.session["bom_query_url"] = query_url
        count = float(bom_query.count())
        pagesize = request.pagesizes if in_page else count
        pagination = Pagination(bom_query, pagesize, page)

        item_id = request.GET.get("id", None)
        child_bom_list = []
        if bom_query:
            if item_id:
                pid = item_id
            else:
                pid = bom_query.first().id
            parent = BomModel.objects.get(id=pid)

            # 是否查询所有数据
            if query_data.get("all", None):
                child_bom = BomModel.objects.filter(parent_id=parent.item.id).order_by("id")
            else:
                child_bom = BomModel.objects.filter(parent_id=parent.item.id,
                                                    effective_end__gte=timezone.now()).order_by("id")
            for i in child_bom:
                adict = {}
                adict["id"] = i.id
                adict["item"] = i.item.nr
                adict["nr"] = i.nr
                adict["effective_start"] = i.effective_start
                adict["effective_end"] = i.effective_end
                adict["qty"] = i.qty
                child_bom_list.append(adict)
            content = {
                "parent_id": pid,
                "parent": parent,
                "pg": pagination,
                "data": node_data,
                "bom_query": bom_query,
                "child_bom_list": child_bom_list,
                "download": download,
                "query": fs,
                "query_url": request.session.get("bom_query_url", None) if request.session.get("bom_query_url",
                                                                                               None) else "",
                "select_op": select_op,
                "select_field": self.select_field,
                "perm": request.perm,
                "error": ""
            }
        else:
            content = {
                "parent_id": "",
                "parent": "",
                "pg": pagination,
                "nodes": pagination.get_objs(),
                "data": node_data,
                "bom_query": bom_query,
                "child_bom_list": child_bom_list,
                "download": download,
                "query": fs,
                "query_url": request.session.get("bom_query_url", None) if request.session.get("bom_query_url",
                                                                                               None) else "",
                "select_op": select_op,
                "select_field": self.select_field,
                "perm": request.perm,
                "error": ""
            }
        return content


class EnumView(View):
    def get(self, request, *args, **kwargs):
        item = int(request.GET.get("item", None))
        child_list = [item]
        parent_data = {}
        data = []
        if item:
            def get_child(bom):
                child = []
                child_node = BomModel.objects.filter(parent_id=bom).order_by("id")
                for i in child_node:
                    child_list.append(i.item_id)
                    child = get_child(i.item_id)
                return child

            # 获取所有当前节点的所有子节点
            get_child(item)
            parent = Item.objects.filter(~Q(id__in=child_list)).order_by("id").values("id", "nr")
            for i in parent:
                parent_data[i["id"]] = i["nr"]
            data = [{"id": k, "nr": v} for k, v in dict(parent_data).items()] if parent_data else None

        return HttpResponse(json.dumps(data, cls=DjangoJSONEncoder),
                            content_type='application/json')


class ItemBomAdd(View):
    items = BomNr.get_nr()

    def get(self, request, *args, **kwargs):
        content = {
            "item": self.items,
        }
        return TemplateResponse(request, "my_app_template/bom_add.html", content)

    def post(self, request, *args, **kwargs):
        form = ItemBomForm(request.POST)
        form_get = request.POST.dict()
        item = form_get["item"]
        parent = form_get["parent"]
        nr = form_get["nr"]
        # 生效开始和生效结束可能为空
        start = datetime(datetime.now().year, datetime.now().month, datetime.now().day)
        end = datetime(2030, 12, 31)
        effective_start = form_get["effective_start"] if form_get["effective_start"] else datetime.strftime(start,
                                                                                                            "%Y-%m-%d")
        effective_end = form_get["effective_end"] if form_get["effective_end"] else datetime.strftime(end, "%Y-%m-%d")
        qty = form_get["qty"]
        # 获取上传的附件
        files = request.FILES.getlist('files', None)
        files_list = []
        imgs_list = []
        content = {
            "item": self.items,
            "parent": self.items,
            "parent_id": int(parent) if parent else "",
            "item_id": int(item),
            "nr": nr,
            "effective_start": effective_start,
            "effective_end": effective_end,
            "qty": qty,
            "error": ""
        }
        # 判断上传文件的格式
        if files:
            for i in files:
                if i.content_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                      'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
                    files_list.append(i)
                elif i.content_type in ['image/png', "image/jpeg"]:
                    imgs_list.append(i)
                else:
                    content["error"] = "请上传Word或者Excel文件"
                    return TemplateResponse(request, "my_app_template/bom_add.html", content)

        if form.is_valid():
            if item == parent:
                content["error"] = "物料和父类信息冲突，请重新填写"
                return TemplateResponse(request, "my_app_template/bom_add.html", content)
            try:
                if parent:
                    add_parent = BomModel.objects.filter(item_id=parent)
                    if not add_parent:
                        content["error"] = "请先创建父类信息"
                        return TemplateResponse(request, "my_app_template/bom_add.html", content)
                    # item是否为根节点
                    add_item = BomModel.objects.filter(item_id=item, parent=None)
                    item_parent = Item.objects.get(id=parent)
                    # 如果是为根节点修改节点信息
                    if add_item:
                        add_item = add_item.first()
                        add_item.parent = item_parent
                        add_item.nr = nr
                        add_item.effective_end = effective_end
                        add_item.effective_start = effective_start
                        add_item.qty = qty
                        add_item.save()
                        content_object = add_item
                    # item如果不是根节点，创建信息
                    else:
                        content_object = BomModel.objects.create(nr=nr, item_id=item, parent_id=parent,
                                                                 effective_end=effective_end,
                                                                 effective_start=effective_start, qty=qty)
                else:
                    content_object = BomModel.objects.create(nr=nr, item_id=item, effective_end=effective_end,
                                                             effective_start=effective_start, qty=qty)

                content_type = ContentType.objects.get(app_label='my_app', model='bommodel')
                if len(imgs_list) > 0:
                    UploadFileModel_img = []
                    for i in imgs_list:
                        img_data = i.read()
                        key_name = "bom" + "-" + nr + "-" + str(uuid.uuid4()) + "-" + i.name
                        try:
                            storage_name = storage(img_data, key_name, mime_type=i.content_type)
                        except Exception as e:
                            print(e)
                            content["error"] = "上传图片到七牛云错误"
                            return TemplateResponse(request, "my_app_template/bom_add.html", content)

                        imgs_add = UploadFileModel(user=request.user, content_type=content_type,
                                                   content_object=content_object,
                                                   img_url=storage_name, img=i.name)
                        UploadFileModel_img.append(imgs_add)

                    UploadFileModel.objects.bulk_create(UploadFileModel_img, batch_size=100)
                if len(files_list) > 0:
                    UploadFileModel_file = []
                    for i in files_list:
                        file_data = i.read()
                        key_name = str(uuid.uuid1()) + i.name
                        try:
                            storage_name = storage(file_data, key_name, mime_type=i.content_type)
                        except Exception as e:
                            content["error"] = "上传文件到七牛云错误"
                            return TemplateResponse(request, "my_app_template/bom_add.html", content)

                        files_add = UploadFileModel(user=request.user, content_type=content_type,
                                                    content_object=content_object,
                                                    file_url=storage_name, file=i.name)
                        UploadFileModel_file.append(files_add)
                    UploadFileModel.objects.bulk_create(UploadFileModel_file, batch_size=100)
            except Exception as e:
                print(e)
                content["error"] = "数据已经存在，请重新填写数据"
                return TemplateResponse(request, "my_app_template/bom_add.html", content)
            else:
                return HttpResponseRedirect("/item/bom/")
        else:
            content["error"] = "表单错误"
            return TemplateResponse(request, "my_app_template/bom_add.html", content)


class ItemBomEdit(View):
    def get(self, request, *args, **kwargs):
        id = request.GET.get("id", None)
        try:
            bom_detail = BomModel.objects.get(id=id)
            content = {
                "id": id,
                "item": bom_detail.item.nr,
                "parent": bom_detail.parent.nr if bom_detail.parent else "",
                "nr": bom_detail.nr,
                "effective_start": datetime.strftime(bom_detail.effective_start, "%Y-%m-%d"),
                "effective_end": datetime.strftime(bom_detail.effective_end, "%Y-%m-%d"),
                "qty": bom_detail.qty,
                "perm": request.perm
            }
            return TemplateResponse(request, "my_app_template/bom_edit.html", content)
        except Exception as e:
            return HttpResponseRedirect("/item/bom/?mag=查询数据失败")

    def post(self, request, *args, **kwargs):
        form = ItemBomForm(request.POST)
        # 获取表单数据
        id = request.POST['id']
        bom_detail = request.POST['item']
        parent = request.POST['parent']
        nr = request.POST['nr']
        effective_start = request.POST["effective_start"] if request.POST["effective_start"] else datetime(
            datetime.now().year,
            datetime.now().month,
            datetime.now().day)
        effective_end = request.POST["effective_end"] if request.POST["effective_end"] else datetime(2030, 12, 31)
        qty = request.POST['qty']
        files = request.FILES.getlist('files', None)
        files_list = []
        imgs_list = []
        content = {
            "id": id,
            "item": bom_detail,
            "parent": parent,
            "nr": nr,
            "effective_start": effective_start,
            "effective_end": effective_end,
            "qty": qty,
            "perm": request.perm
        }
        # 判断上传文件的格式
        if files:
            for i in files:
                if i.content_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                      'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
                    files_list.append(i)
                elif i.content_type in ['image/png', "image/jpeg"]:
                    imgs_list.append(i)
                else:
                    content["error"] = "请上传Word或者Excel文件"
                    return TemplateResponse(request, "my_app_template/bom_add.html", content)

        if form.is_valid():

            with transaction.atomic(savepoint=False):
                # 创建保存点
                try:
                    bom_detail = BomModel.objects.get(id=id)
                    bom_detail.effective_start = effective_start
                    bom_detail.effective_end = effective_end
                    bom_detail.qty = qty
                    bom_detail.nr = nr
                    bom_detail.save()
                    content_type = ContentType.objects.get(app_label='my_app', model='bommodel')
                    before_upload = UploadFileModel.objects.filter(content_type=content_type, object_pk=id)
                    # 删除上传文件
                    for i in before_upload:
                        try:
                            if i.file:
                                delete_qiniu_file(i.file_url)
                            if i.img:
                                delete_qiniu_file(i._url)
                        except Exception as e:
                            print(e)
                            content["error"] = "删除之前文件错误"
                            return TemplateResponse(request, "my_app_template/bom_edit.html", content)

                    before_upload.delete()
                    if len(imgs_list) > 0:
                        UploadFileModel_img = []
                        for i in imgs_list:
                            img_data = i.read()
                            key_name = "bom" + "-" + nr + "-" + str(uuid.uuid4()) + "-" + i.name
                            try:
                                storage_name = storage(img_data, key_name, mime_type=i.content_type)
                            except Exception as e:
                                print(e)
                                content["error"] = "上传图片到七牛云错误"
                                return TemplateResponse(request, "my_app_template/bom_edit.html", content)

                            imgs_add = UploadFileModel(user=request.user, content_type=content_type,
                                                       content_object=bom_detail,
                                                       img_url=storage_name, img=i.name)
                            UploadFileModel_img.append(imgs_add)

                        UploadFileModel.objects.bulk_create(UploadFileModel_img, batch_size=100)
                    if len(files_list) > 0:
                        UploadFileModel_file = []
                        for i in files_list:
                            file_data = i.read()
                            key_name = "bom" + "-" + nr + "-" + str(uuid.uuid4()) + "-" + i.name
                            try:
                                storage_name = storage(file_data, key_name, mime_type=i.content_type)
                            except Exception as e:
                                print(e)
                                content["error"] = "上传文件到七牛云错误"
                                return TemplateResponse(request, "my_app_template/bom_edit.html", content)

                            files_add = UploadFileModel(user=request.user, content_type=content_type,
                                                        content_object=bom_detail,
                                                        file_url=storage_name, file=i.name)
                            UploadFileModel_file.append(files_add)

                        UploadFileModel.objects.bulk_create(UploadFileModel_file, batch_size=100)
                except Exception as e:
                    print(e)
                    content["error"] = "编辑失效"
                    return TemplateResponse(request, "my_app_template/bom_edit.html", content)
                else:
                    return HttpResponseRedirect("/item/bom/")
        else:
            content["error"] = "表单填写错误"
            return TemplateResponse(request, "my_app_template/bom_edit.html", content)


class ItemBomCalculate(View):
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('format', None)
        if fmt is None:
            content = self._get_data(request)
            return TemplateResponse(request, "my_app_template/bom_calculate.html", content)
        else:
            # 下载Excel
            data = self._get_data(request)
            wb = Workbook(write_only=True)
            title = "bom计算"
            ws1 = wb.create_sheet("采购物料")
            ws2 = wb.create_sheet("制造物料")
            header_ws1 = []
            header_ws2 = []
            header1 = []
            header2 = []

            common_headers = ("物料", "物料数量")
            purchase_headers = ("采购物料", "数量")
            manufacture_headers = ("制造物料", "数量")
            body_fields = ("nr", "qty")

            for i in common_headers:
                cell = WriteOnlyCell(ws1, value=i)
                header_ws1.append(cell)
            ws1.append(header_ws1)
            cell1 = WriteOnlyCell(ws1, value=data["item"])
            cell2 = WriteOnlyCell(ws1, value=data["qty"])
            body_ws1 = [cell1, cell2]
            ws1.append(body_ws1)

            if data['purchase']:
                for h in purchase_headers:
                    cell = WriteOnlyCell(ws1, value=h)
                    header1.append(cell)
                ws1.append(header1)
                for i in data['purchase']:
                    body = []
                    for field in body_fields:
                        if field in i:
                            cell = WriteOnlyCell(ws1, value=i[field])
                            body.append(cell)
                    ws1.append(body)

            for i in common_headers:
                cell = WriteOnlyCell(ws2, value=i)
                header_ws2.append(cell)
            ws2.append(header_ws2)
            cell1 = WriteOnlyCell(ws2, value=data["item"])
            cell2 = WriteOnlyCell(ws2, value=data["qty"])
            body_ws2 = [cell1, cell2]
            ws2.append(body_ws2)

            if data['manufacture']:
                for h in manufacture_headers:
                    cell = WriteOnlyCell(ws2, value=h)
                    header2.append(cell)
                ws2.append(header2)
                for i in data['manufacture']:
                    body = []
                    for field in body_fields:
                        if field in i:
                            cell = WriteOnlyCell(ws2, value=i[field])
                            body.append(cell)
                    ws2.append(body)
            output = BytesIO()
            wb.save(output)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                content=output.getvalue()
            )
            response['Content-Disposition'] = "attachment; filename*=utf-8''%s.xlsx" % urllib.parse.quote(
                force_str(title))
            response['Cache-Control'] = "no-cache, no-store"
            return response

    def _get_data(self, request, *args, **kwargs):
        id = request.GET.get("id", None)
        number = request.GET.get("qty", None)
        if number:
            number = int(number)
        bom_filter = BomModel.objects.filter(id=id)
        if not bom_filter:
            return HttpResponseRedirect("/item/bom/?msg=查询数据出错")
        bom = bom_filter.first()
        # 计算采购物料的qty
        purchase_list = []
        manufacture_list = []

        def get_qty(bom, qty):
            qty_child = qty
            for i in BomModel.objects.filter(parent_id=bom.item.id, effective_end__gte=timezone.now()).order_by("id"):
                qty_math = i.qty * qty
                adict = {"nr": i.item.nr, "qty": qty_math}
                node = BomModel.objects.filter(parent_id=i.item.id, effective_end__gte=timezone.now())
                # 判断是否是叶子节点
                if not node:
                    # 采购物料
                    purchase_list.append(adict)
                    qty_child = get_qty(i, qty_math)
                else:
                    # 制造物料
                    manufacture_list.append(adict)
                    qty_child = get_qty(i, qty_math)

            return qty_child

        get_qty(bom, number / bom.qty)

        content = {
            "id": id,
            "item": bom.item.nr,
            "qty": number,
            "purchase": purchase_list,
            "manufacture": manufacture_list,
            "error": ""
        }
        return content


class ItemBomDelete(View):
    def get(self, request, *args, **kwargs):
        id = request.GET.get("id", None)
        if id:
            try:
                bom_detail = BomModel.objects.get(id=id)
            except:
                return HttpResponseRedirect("/bom/?msg=找不到此数据")
            else:
                content = {
                    "id": id,
                    "item": bom_detail.item.nr,
                    "parent": bom_detail.parent.nr if bom_detail.parent else "",
                    "nr": bom_detail.nr,
                    "effective_start": datetime.strftime(bom_detail.effective_start, "%Y-%m-%d"),
                    "effective_end": datetime.strftime(bom_detail.effective_start, "%Y-%m-%d"),
                    "qty": bom_detail.qty
                }
                return TemplateResponse(request, "my_app_template/bom_delete.html", content)
        else:
            return HttpResponseRedirect("/item/bom?msg=找不到数据")

    def post(self, request, *args, **kwargs):
        id = request.POST['id']
        parent = request.POST['parent']
        item = request.POST['item']
        nr = request.POST['nr']
        effective_start = request.POST['effective_start']
        effective_end = request.POST['effective_end']
        qty = request.POST["qty"]

        content = {
            "id": id,
            "item": item,
            "parent": parent,
            "nr": nr,
            "effective_start": effective_start,
            "effective_end": effective_end,
            "qty": qty,
            "perm": request.perm,
        }

        with transaction.atomic(savepoint=False):
            # 创建保存点
            try:
                bom = BomModel.objects.get(id=id)
                bom.delete()
            except Exception as e:
                traceback.print_exc()
                content["error"] = "删除失败"
                return TemplateResponse(request, "my_app_template/bom_delete.html", content)
            else:
                return HttpResponseRedirect('/item/bom/')


class UploadView(View):
    def get(self, request, *args, **kwargs):
        content_parameter = request.GET['content_parameter']
        pid = request.GET['id']
        content_type = ContentType.objects.filter(app_label='my_app', model=content_parameter.lower())
        upload_list = []
        if content_type:
            upload = UploadFileModel.objects.filter(content_type=content_type.first(), object_pk=pid).order_by('-id')
            for i in upload:
                adict = {}
                adict["file_url"] = QINIU_DOMIN_PREFIX + i.file_url if i.file_url else None
                adict["file"] = i.file
                adict["img_url"] = QINIU_DOMIN_PREFIX + i.img_url if i.img_url else None
                adict["img"] = i.img
                adict["user"] = i.user.username
                adict["upload_time"] = i.upload_time
                upload_list.append(adict)
            return TemplateResponse(request, "my_app_template/bom_files.html", {"content": upload_list})
        else:
            return TemplateResponse(request, "my_app_template/bom_files.html", {"error": "参数错误"})


class JustTest(View):
    def get(self, request, *args, **kwargs):
        # url_string = "http://pl3u05m5t.bkt.clouddn.com/bom-AC-b67967ab-eedd-4bb5-a5f5-cd9cef883bca-admin%E5%8A%B3%E9%94%90%E5%BC%80%E7%A5%A8%E4%BF%A1%E6%81%AF%E8%B5%84%E6%96%99%20%281%29.docx"
        # r = requests.get(url_string, stream=True)
        # content = r.text
        # # content = content.decode("utf-8")
        # # return HttpResponse(content)
        # response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = 'attachment; filename="somefilename.pdf"'
        # p = canvas.Canvas(response)
        # p.drawString(500, 0, content)
        # p.showPage()
        # p.save()
        logging.error("test~~~~~~")
        return HttpResponse("ok")



