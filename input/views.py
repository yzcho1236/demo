import traceback
import urllib
from datetime import datetime
from io import BytesIO

from django.db import transaction
from django.db.models import NOT_PROVIDED, AutoField
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.utils.encoding import force_str
from django.views import View
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from django.utils import timezone
from demo.util import Pagination, GetBom, GetQueryData, select_op
from input.form import BomForm, BomEditForm
from input.models import Tree_Model


class Bom(View):
    file_headers = ('id', '物料', '物料代码', '组成物料', '组成物料代码', '生效开始', "生效结束", "数量")
    select_field = {"id": "id", 'parent__name': '物料', "parent__nr": '物料代码', 'name': '组成物料', 'nr': '组成物料代码',
                    'effective_start': '生效开始',
                    'effective_end': '生效结束'}
    uploader_header = {"id": "id", "物料": "parent", "物料代码": "parent_nr", "组成物料": "name", "组成物料代码": "nr",
                       "生效开始": "effective_start", "生效结束": "effective_end",
                       "数量": "qty"}

    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('format', None)
        if fmt is None:
            content = self._get_data(request)
            return TemplateResponse(request, "bom.html", content)
        elif fmt == 'spreadsheet':
            # 下载 excel
            json = self._get_data(request, in_page=False)
            wb = Workbook(write_only=True)
            title = "物料清单"
            ws = wb.create_sheet(title)

            headers = []
            for h in self.file_headers:
                cell = WriteOnlyCell(ws, value=h)
                headers.append(cell)
            ws.append(headers)

            body_fields = ("id", "parent", "parent_nr", "item", "item_nr", "effective_start", "effective_end", "qty")
            data = []
            for i in json["nodes"]:
                adict = {}
                adict["id"] = i.id
                adict["item"] = i.name
                adict["item_nr"] = i.nr
                if i.parent:
                    adict["parent"] = i.parent.name
                    adict["parent_nr"] = i.parent.nr
                else:
                    adict["parent"] = ""
                    adict["parent_nr"] = ""
                adict["effective_start"] = i.effective_start
                adict["effective_end"] = i.effective_end
                adict["qty"] = i.qty
                data.append(adict)

            for i in data:
                body = []
                for field in body_fields:
                    if field in i:
                        cell = WriteOnlyCell(ws, value=i[field])
                        body.append(cell)
                ws.append(body)
            output = BytesIO()
            wb.save(output)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                content=output.getvalue()
            )
            response['Content-Disposition'] = "attachment; filename*=utf-8''%s.xlsx" % urllib.parse.quote(
                force_str(title))
            response['Cache-Control'] = "no-cache, no-store"
            return response

    def post(self, request, *args, **kwargs):
        """上传Excel"""

        data = self._get_data(request)
        if request.FILES and len(request.FILES) == 1:
            count = 0
            for filename, file in request.FILES.items():
                if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    count += 1
            if count == 0:
                data["error"] = "请选择Excel文件进行上传"
                return TemplateResponse(request, "item.html", data)
            else:
                for name, file in request.FILES.items():
                    wb = load_workbook(filename=file, read_only=True, data_only=True)
                    # 第一个sheet
                    sheet = wb[wb.sheetnames[0]]
                    row_count = 0
                    row_number = 1
                    # 对应表头是在第几列
                    headers_index = {}
                    # Excel传入的字段和模型类字段的对应关系
                    headers_field_name = {}
                    excel_level = 0
                    for row in sheet.iter_rows():
                        excel_level += 1
                        row_count += 1
                        if row_number == 1:
                            row_number += 1
                            # 获取excelheader
                            _headers = [i.value for i in row]
                            index = 0

                            for h in _headers:
                                if h in self.uploader_header:
                                    headers_index[h] = index
                                    headers_field_name[h] = self.uploader_header[h]
                                index += 1
                        # 取值
                        else:
                            values = [i.value for i in row]
                            none_len = 0
                            for v in values:
                                if v is None:
                                    none_len += 1

                            if none_len == len(values):
                                continue
                            upload_fields = [v for k, v in headers_field_name.items()]

                            # 必传字段
                            required_fields = set()
                            for i in Tree_Model._meta.fields:
                                if not i.blank and i.default == NOT_PROVIDED and not isinstance(i, AutoField):
                                    required_fields.add(i.name)
                            aset = {"level", "rght", "tree_id", "lft"}
                            required_fields = required_fields.difference(aset)
                            required_fields.update(["parent", "parent_nr"])
                            set_header = set(upload_fields)
                            common_fields = set_header.intersection(required_fields)
                            if common_fields != required_fields:
                                data["error"] = "上传字段不全，请重新上传"
                                return TemplateResponse(request, "bom.html", data)
                            adict = {}
                            for k, v in headers_index.items():
                                value = values[v]
                                field_name = headers_field_name[k]
                                if field_name == "name":
                                    adict["name"] = value
                                if field_name == "nr":
                                    adict["nr"] = value
                                if field_name == "qty":
                                    adict["qty"] = value
                                if field_name == "effective_start":
                                    adict["effective_start"] = value if value else datetime(datetime.now().year,
                                                                                            datetime.now().month,
                                                                                            datetime.now().day)
                                if field_name == "effective_end":
                                    adict["effective_end"] = value if value else datetime(2030, 12, 31)

                                if field_name == "parent":
                                    adict["parent"] = value
                                if field_name == "parent_nr":
                                    adict["parent_nr"] = value

                            for k, v in adict.items():
                                if k in ["name", "qty", "nr"] and v is None:
                                    data["error"] = "%s 字段值不能为空" % k
                                    return TemplateResponse(request, "bom.html", data)

                            if adict["name"] and adict["nr"]:
                                item_node = Tree_Model.objects.filter(name=adict["name"], nr=adict["nr"])
                                if item_node:
                                    item_node = item_node.first()
                                    alist = list(filter(lambda x: x, [adict["parent_nr"], adict["parent"]]))
                                    if len(alist) == 1:
                                        data["error"] = "第 %s 行请填写完整的物料和物料代码" % excel_level
                                        return TemplateResponse(request, "bom.html", data)
                                    elif len(alist) == 2:
                                        parent_node = Tree_Model.objects.filter(name=adict["parent"],
                                                                                nr=adict["parent_nr"])
                                        item_parent_node = Tree_Model.objects.filter(name=adict["name"],
                                                                                     nr=adict["nr"],
                                                                                     parent__nr=adict["parent_nr"],
                                                                                     parent__name=adict["parent"])
                                        excel_node = Tree_Model.objects.filter(name=adict["name"],
                                                                               parent__nr=adict["parent_nr"],
                                                                               parent__name=adict["parent"])
                                        if not parent_node:
                                            data["error"] = "第 %s 行信息填写错误" % excel_level
                                            return TemplateResponse(request, "bom.html", data)
                                        else:
                                            # 如果物料信息匹配正确，更新
                                            if item_parent_node:
                                                item_node.effective_start = adict["effective_start"]
                                                item_node.effective_end = adict["effective_end"]
                                                item_node.qty = adict["qty"]
                                                item_node.save()
                                            # 如果没有item_parent_node 添加
                                            elif not item_parent_node and not excel_node:
                                                with transaction.atomic(savepoint=False):
                                                    try:

                                                        parent_nodes = Tree_Model.objects.filter(name=adict["parent"])
                                                        create_list = []
                                                        for i in parent_nodes:
                                                            # TODO 创建时没有nr值
                                                            nr = i.name + "-" + adict["nr"] + "-" + str(i.level)
                                                            create_nodes = Tree_Model.objects.create(name=adict["name"],
                                                                                                     parent=i, nr=nr,
                                                                                                     effective_start=
                                                                                                     adict[
                                                                                                         "effective_start"],
                                                                                                     effective_end=
                                                                                                     adict[
                                                                                                         "effective_end"],
                                                                                                     qty=adict["qty"])
                                                            create_list.append(create_nodes)

                                                        def get_child(item, parent_node):
                                                            level = item.level
                                                            for i in item.get_children().filter(
                                                                    effective_end__gte=timezone.now()):
                                                                nr = i.name + "-" + adict["nr"] + "-" + str(level)
                                                                Tree_Model.objects.create(name=i.name, nr=nr,
                                                                                          parent=parent_node,
                                                                                          effective_start=i.effective_start,
                                                                                          effective_end=i.effective_end,
                                                                                          qty=i.qty)
                                                                level = get_child(i, item)
                                                            return level

                                                        for i in create_list:
                                                            get_child(item_node, i)

                                                    except Exception as e:
                                                        print(e)
                                                        traceback.print_exc()
                                                        data["error"] = e
                                                        return TemplateResponse(request, "bom.html", data)
                                            elif not item_parent_node and excel_node:
                                                data["error"] = "第 %s 行填写正确的物料名称、代码信息" % excel_level
                                                return TemplateResponse(request, "bom.html", data)

                                    # 如果没有父类信息，更新
                                    elif len(alist) == 0:
                                        if not item_node:
                                            data["error"] = "第 %s 行组成物料不存在" % excel_level
                                            return TemplateResponse(request, "bom.html", data)
                                        else:
                                            item_node.effective_start = adict["effective_start"]
                                            item_node.effective_end = adict["effective_end"]
                                            item_node.qty = adict["qty"]

                            else:
                                data["error"] = "第 %s 行请填写完整的组成物料信息，包含名称、代码" % excel_level
                                return TemplateResponse(request, "bom.html", data)
                query_url = request.session["bom_query_url"]
                return HttpResponseRedirect("/bom/?page=1" + query_url)
        else:
            data["error"] = "没有上传的文件"
            return TemplateResponse(request, "bom.html", data)

    def _get_data(self, request, in_page=True, *args, **kwargs):
        # 获取到的页面数
        page = request.GET.get("page", 1)
        query_data = dict(request.GET.lists())
        item_query, fs, query_url, search = GetQueryData.get_tree(Tree_Model, query_data, self.select_field)
        if search:
            request.session["bom_query_url"] = query_url
        count = float(item_query.count())
        pagesize = 20 if in_page else count
        pagination = Pagination(item_query, pagesize, page)

        item_id = request.GET.get("id", None)
        page = request.GET.get("page", 1)
        data = []
        if item_query:
            if item_id:
                id = item_id
            else:
                id = item_query.first().id
            parent = Tree_Model.objects.get(id=id)
            pagination = Pagination(item_query, request.pagesizes, page)
            items = Tree_Model.objects.filter(parent__id=id, effective_end__gte=timezone.now())
            for i in items:
                adict = {}
                adict["id"] = i.id
                adict["item"] = i.name
                adict["item_nr"] = i.nr
                adict["parent"] = parent.name
                adict["parent_nr"] = parent.nr
                adict["effective_start"] = i.effective_start
                adict["effective_end"] = i.effective_end
                adict["qty"] = i.qty
                data.append(adict)
            content = {
                "id": id,
                "parent": parent.name,
                "parent_nr": parent.nr,
                "pg": pagination,
                "nodes": pagination.get_objs(),
                "data": data,
                "query": fs,
                "query_url": request.session.get("bom_query_url", None) if request.session.get("bom_query_url",
                                                                                               None) else "",
                "select_op": select_op,
                "select_field": self.select_field,
                "perm": request.perm,
                "error": ""
            }
        else:
            content = {
                "id": "",
                "parent": "",
                "pg": pagination,
                "nodes": pagination.get_objs(),
                "data": data,
                "query": fs,
                "query_url": request.session.get("bom_query_url", None) if request.session.get("bom_query_url",
                                                                                               None) else "",
                "select_op": select_op,
                "select_field": self.select_field,
                "perm": request.perm,
                "error": ""
            }
        return content


class BomAdd(View):
    def get(self, request, *args, **kwargs):
        parent = request.GET.get("name", None)
        parent_nr = request.GET.get("nr", None)
        parent_id = request.GET.get("id", None)
        items = GetBom.get_bom(parent_id, parent)
        content = {
            "parent_id": parent_id,
            "parent": parent,
            "parent_nr": parent_nr,
            "items": items,
            "perm": request.perm
        }
        return TemplateResponse(request, "bom_add.html", content)

    def post(self, request, *args, **kwargs):
        form = BomForm(request.POST)
        form_get = request.POST.dict()
        parent_id = form_get["parent_id"]
        parent_nr = form_get["parent_nr"]
        parent = form_get["parent"]
        item = form_get["item"]
        item_nr = form_get["item_nr"]
        # 生效开始和生效结束可能为空
        effective_start = form_get["effective_start"] if form_get["effective_start"] else datetime(datetime.now().year,
                                                                                                   datetime.now().month,
                                                                                                   datetime.now().day)
        effective_end = form_get["effective_end"] if form_get["effective_end"] else datetime(2030, 12, 31)
        qty = form_get["qty"]
        content = {
            "parent_id": parent_id,
            "parent": parent,
            "parent_nr": parent_nr,
            "items": GetBom.get_bom(parent_id, parent),
            "item": item,
            "item_nr": item_nr,
            "effective_start": effective_start,
            "effective_end": effective_end,
            "qty": qty,
            "error": ""
        }
        if form.is_valid():
            if Tree_Model.objects.filter(nr=item_nr):
                content["error"] = "组成物料代码已存在，请重新输入"
                return TemplateResponse(request, "bom_add.html", content)
            with transaction.atomic(savepoint=False):
                try:
                    child = Tree_Model.objects.filter(name=item, effective_end__gte=timezone.now()).first()

                    # 如在W节点下面添加Q，查询所有的W，然后全部添加Q
                    # 查询所有的parent
                    parent_nodes = Tree_Model.objects.filter(name=parent)

                    # 创建(批量创建会报错)

                    # create_list = []
                    # for i in parent_nodes:
                    #     create_node = Tree_Model(name=item, parent=i, effective_start=effective_start,
                    #                              effective_end=effective_end, qty=qty)
                    #     create_list.append(create_node)
                    #
                    # create_nodes = Tree_Model.objects.bulk_create(create_list)

                    create_list = []
                    for i in parent_nodes:
                        # TODO 创建时没有nr值
                        nr = i.name + "-" + item_nr + "-" + str(i.level)
                        create_nodes = Tree_Model.objects.create(name=item, parent=i, nr=nr,
                                                                 effective_start=effective_start,
                                                                 effective_end=effective_end, qty=qty)
                        create_list.append(create_nodes)

                    def get_child(item, parent_node):
                        level = item.level
                        for i in item.get_children().filter(effective_end__gte=timezone.now()):
                            nr = i.name + "-" + item_nr + "-" + str(level)
                            Tree_Model.objects.create(name=i.name, nr=nr, parent=parent_node,
                                                      effective_start=i.effective_start,
                                                      effective_end=i.effective_end, qty=i.qty)
                            level = get_child(i, item)
                        return level

                    if child:
                        for i in create_list:
                            get_child(child, i)

                except Exception as e:
                    print(e)
                    traceback.print_exc()
                    content["error"] = e
                    return TemplateResponse(request, "bom_add.html", content)
                else:
                    return HttpResponseRedirect("/bom/")
        else:
            content["error"] = "表单错误"
            return TemplateResponse(request, "bom_add.html", content)


class BomEdit(View):
    def get(self, request, *args, **kwargs):
        id = request.GET.get("id", None)
        if id:
            item = Tree_Model.objects.get(id=id)
            content = {
                "id": id,
                "name": item.name,
                "effective_start": datetime.strftime(item.effective_start, "%Y-%m-%d"),
                "effective_end": datetime.strftime(item.effective_end, "%Y-%m-%d"),
                "qty": item.qty,
                "perm": request.perm
            }
            return TemplateResponse(request, "bom_edit.html", content)
        else:
            return HttpResponseRedirect("/bom/?mag=查询数据失败")

    def post(self, request, *args, **kwargs):
        form = BomEditForm(request.POST)
        # 获取表单数据
        id = request.POST['id']
        name = request.POST['name']
        effective_start = request.POST["effective_start"] if request.POST["effective_start"] else datetime(
            datetime.now().year,
            datetime.now().month,
            datetime.now().day)
        effective_end = request.POST["effective_end"] if request.POST["effective_end"] else datetime(2030, 12, 31)
        qty = request.POST['qty']
        content = {
            "id": id,
            "name": name,
            "effective_start": effective_start,
            "effective_end": effective_end,
            "qty": qty,
            "perm": request.perm
        }
        if form.is_valid():

            with transaction.atomic(savepoint=False):
                # 创建保存点
                try:
                    item = Tree_Model.objects.get(id=id)
                    item.name = name
                    item.effective_start = effective_start
                    item.effective_end = effective_end
                    item.qty = qty
                    item.save()
                except Exception as e:
                    print(e)
                    content["error"] = "编辑失效"
                    return TemplateResponse(request, "bom_edit.html", content)
                else:
                    return HttpResponseRedirect("/bom/")
        else:
            content["error"] = "表单填写错误"
            return TemplateResponse(request, "bom_edit.html", content)


class BomCalculate(View):
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('format', None)
        if fmt is None:
            content = self._get_data(request)
            return TemplateResponse(request, "bom_calculate.html", content)
        else:
            # 下载Excel
            data = self._get_data(request)
            wb = Workbook(write_only=True)
            title = "bom计算"
            ws1 = wb.create_sheet("采购物料")
            ws2 = wb.create_sheet("制造物料")
            header_ws1 = []
            header_ws2 = []
            header1 = []
            header2 = []

            common_headers = ("物料", "物料数量")
            purchase_headers = ("采购物料", "数量")
            manufacture_headers = ("制造物料", "数量")
            body_fields = ("name", "qty")

            for i in common_headers:
                cell = WriteOnlyCell(ws1, value=i)
                header_ws1.append(cell)
            ws1.append(header_ws1)
            cell1 = WriteOnlyCell(ws1, value=data["item"])
            cell2 = WriteOnlyCell(ws1, value=data["qty"])
            body_ws1 = [cell1, cell2]
            ws1.append(body_ws1)

            if data['purchase']:
                for h in purchase_headers:
                    cell = WriteOnlyCell(ws1, value=h)
                    header1.append(cell)
                ws1.append(header1)
                for i in data['purchase']:
                    body = []
                    for field in body_fields:
                        if field in i:
                            cell = WriteOnlyCell(ws1, value=i[field])
                            body.append(cell)
                    ws1.append(body)

            for i in common_headers:
                cell = WriteOnlyCell(ws2, value=i)
                header_ws2.append(cell)
            ws2.append(header_ws2)
            cell1 = WriteOnlyCell(ws2, value=data["item"])
            cell2 = WriteOnlyCell(ws2, value=data["qty"])
            body_ws2 = [cell1, cell2]
            ws2.append(body_ws2)

            if data['manufacture']:
                for h in manufacture_headers:
                    cell = WriteOnlyCell(ws2, value=h)
                    header2.append(cell)
                ws2.append(header2)
                for i in data['manufacture']:
                    body = []
                    for field in body_fields:
                        if field in i:
                            cell = WriteOnlyCell(ws2, value=i[field])
                            body.append(cell)
                    ws2.append(body)
            output = BytesIO()
            wb.save(output)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                content=output.getvalue()
            )
            response['Content-Disposition'] = "attachment; filename*=utf-8''%s.xlsx" % urllib.parse.quote(
                force_str(title))
            response['Cache-Control'] = "no-cache, no-store"
            return response

    def _get_data(self, request, *args, **kwargs):
        id = request.GET.get("id", None)
        number = int(request.GET.get("qty", None))
        item = Tree_Model.objects.get(id=id)

        # 计算采购物料的qty
        purchase_list = []
        manufacture_list = []
        purchase = item.get_leafnodes().filter(effective_end__gte=timezone.now())
        child = item.get_descendants().filter(effective_end__gte=timezone.now())

        if list(purchase) == list(child):
            for i in purchase:
                adict = {}
                adict["qty"] = i.qty * number * item.qty
                adict["name"] = i.name
                purchase_list.append(adict)

        else:
            def get_qty(item, qty):
                qty_child = qty
                for i in item.get_children().filter(effective_end__gte=timezone.now()):
                    qty_math = i.qty * qty
                    adict = {"name": i.name, "qty": qty_math}
                    if i.is_leaf_node():
                        purchase_list.append(adict)
                        qty_child = get_qty(i, qty_math)
                    else:
                        manufacture_list.append(adict)
                        qty_child = get_qty(i, qty_math)

                return qty_child

            get_qty(item, item.qty * number)
            print(purchase_list)
            print(manufacture_list)

        content = {
            "id": id,
            "item": item.name,
            "qty": number,
            "purchase": purchase_list,
            "manufacture": manufacture_list
        }
        return content


class BomDelete(View):
    def get(self, request, *args, **kwargs):
        id = request.GET.get("id", None)
        if id:
            try:
                item = Tree_Model.objects.get(id=id)
            except:
                return HttpResponseRedirect("/bom/?msg=找不到此数据")
            else:
                content = {
                    "id": id,
                    "item": item.name,
                    "parent": item.parent.name if item.parent else "",
                    "effective_start": datetime.strftime(item.effective_start, "%Y-%m-%d"),
                    "effective_end": datetime.strftime(item.effective_start, "%Y-%m-%d"),
                    "qty": item.qty
                }
                return TemplateResponse(request, "bom_delete.html", content)
        else:
            return HttpResponseRedirect("/bom?msg=找不到数据")

    def post(self, request, *args, **kwargs):
        id = request.POST['id']
        parent = request.POST['parent']
        item = request.POST['item']
        effective_start = request.POST['effective_start']
        effective_end = request.POST['effective_end']
        qty = request.POST["qty"]

        content = {
            "id": id,
            "item": item,
            "parent": parent,
            "effective_start": effective_start,
            "effective_end": effective_end,
            "qty": qty,
            "perm": request.perm,
        }

        with transaction.atomic(savepoint=False):
            # 创建保存点
            try:
                item_node = Tree_Model.objects.get(id=id)
                # 判断当前删除的节点是否为根节点，如果为根节点直接删除当前节点
                if not item_node.is_root_node():
                    nodes = Tree_Model.objects.filter(name=item)
                    root = False
                    # 当前节点不是根节点，但是存在是根节点的关系，删除当前节点
                    for i in nodes:
                        if i.is_root_node():
                            root = True
                            break
                    if root:
                        item_node.delete()
                    # 当前节点不存在是根节点的情况，删除所有
                    else:
                        nodes.delete()
                else:
                    item_node.delete()

            except Exception as e:
                traceback.print_exc()
                content["error"] = "删除失败"
                return TemplateResponse(request, "bom_delete.html", content)

            else:
                return HttpResponseRedirect('/bom/')
